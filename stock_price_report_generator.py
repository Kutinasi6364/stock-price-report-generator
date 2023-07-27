# 株価と配当率のスクレイピング
import csv
import time
import datetime
import openpyxl

from bs4 import BeautifulSoup
import requests
import urllib.robotparser


# 各URLの設定
robot_url = "https://minkabu.jp/robots.txt"
basic_url = "https://minkabu.jp/stock/"

# 証券番号のリストを作成
code_list = [
    "8316",
    "9436",
    "8424",
    "8306",
    "4502",
    "9986",
    "7995",
    "8898",
    "7751",
    "8591",
    "8053",
    "8425",
    "8566",
    "9433",
    "2169",
    "4762",
    "7820"
]

#保存先ファイル・シート名
file_path = r"C:\Users\kutin\Desktop\投資\ポートフォリオ管理表(自動).xlsx"
sheet_name = "日本株管理表"

item = []

# クローリング可能かチェック
def check_crawling(url):
    ur = urllib.robotparser.RobotFileParser()
    ur.set_url(robot_url)
    ur.read()
    result = ur.can_fetch("*", url)
    return result

# スクレイピングの実行
def scraping(url, code):
    # 証券番号を最初に追加する
    item_list = [code]

    response = requests.get(url)
    soup = BeautifulSoup(response.content, "html.parser")

    # 配当金抽出
    item_value = soup.find_all(
        "td", attrs={"class": "ly_vamd_inner ly_colsize_9_fix fwb tar wsnw"}
    )
    # 株価抽出
    item_price = soup.find("div", attrs={"class": "stock_price"}).text

    # 抽出した文字の整形
    item_price = item_price.replace(",", "")
    item_price = item_price.replace("\n", "")
    item_price = item_price.replace(" ", "")
    item_price = item_price.replace("円", "")

    # 株価と配当金をリストに追加
    if item_price == "---":
        item_list.append(0)
    else:
        item_list.append(float(item_price))
    
    if item_value[4].string == "---": # 沖縄セルラーの値が取れなかったため対応
        item_list.append(0.01)
    else:
        item_list.append(item_value[4].string.replace("%", ""))
        
    return item_list


# CSVファイルに書き込み
def create_csv(item):
    # 項目名の設定
    # list = ["証券コード", "株価", "配当利回り"]
    with open("stock_price.csv", "w", newline="") as cf:
        cf = csv.writer(cf, delimiter=",")
        #cf.writerow(item)
        for w in item:
            cf.writerow(w)

# csvをExcelファイルに出力
def invet_xls():
    date = datetime.datetime.now()

    # ファイルを開く
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_name]

    # CSVファイルからExcelファイルへ書き込み
    with open("stock_price.csv", "r") as r:
        r = csv.reader(r, delimiter=",")
        #インデックス番号4から開始
        for i, row in enumerate(r, 4):
            sheet.cell(i, 3, value=float(row[2])/100).number_format = "0.00%"
            sheet.cell(i, 4, value=float(row[1])).number_format = "#,###"
        sheet["F1"].value = date.strftime("%Y年%m月%d日")

    # Excelファイルの保存
    book.save(file_path)

# 全ての証券番号の情報を抜き出す
for code in code_list:
    # 指定した証券番号のURLを作成
    url = basic_url + code
    if check_crawling(url):
        # 証券番号、株価、配当金をリストに追加
        item.append(scraping(url, code))
        print(code + "抽出完了")
        time.sleep(2)
    else:
        print("このページはクローリング禁止です")

# CSVファイルの作成→Excel出力
create_csv(item)
invet_xls()
print("出力完了")


