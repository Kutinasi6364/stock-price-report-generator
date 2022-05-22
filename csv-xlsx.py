import csv
import datetime
import openpyxl

file_path = "/home/kutinasi6364/ポートフォリオ管理表(自動).xlsx"
sheet_name = "日本株管理表"

date = datetime.datetime.now()

# ファイルを開く
book = openpyxl.load_workbook(file_path)
sheet = book[sheet_name]

# CSVファイルからExcelファイルへ書き込み
with open("investment.csv", "r") as r:
    r = csv.reader(r, delimiter=",")
    for i, row in enumerate(r, 4):
        sheet.cell(i, 3, value=row[2])
        sheet.cell(i, 4, value=row[1])
    sheet["F1"].value = date.strftime("%Y年%m月%d日")

# Excelファイルの保存
book.save(file_path)
