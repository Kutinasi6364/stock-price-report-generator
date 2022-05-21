# ライブラリのインポート
import openpyxl
import csv

new_book = openpyxl.Workbook()
active_sheet = new_book.active
active_sheet.title = "titleurl"

with open("title-url.csv", "r") as tu:
    tu = csv.reader(tu, delimiter=",")
    for i, row in enumerate(tu, 1):
        active_sheet.cell(i, 1, value=row[0])
        active_sheet.cell(i, 2, value=row[1])

new_book.save("articletitle.xlsx")
